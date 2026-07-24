"""
extract.py
----------
Kéo dữ liệu video YouTube theo từ khóa, có filter khoảng thời gian đăng video.
Lấy đủ các field theo README của project Youtube-ETL_Project, cộng thêm Channel ID + Subscriber Count:
    Video ID, Video Title, Channel ID, Channel Name, Subscriber Count, Publish Date,
    Video URL, Duration, View Count, Like Count, Comment Count, Tags, Category,
    Thumbnail URL, Collection Date

Yêu cầu:
    pip install requests openpyxl python-dotenv

Cách lấy API Key:
    1. Vào https://console.cloud.google.com/
    2. Tạo project mới (hoặc dùng project có sẵn)
    3. Bật "YouTube Data API v3"
    4. Vào Credentials > Create Credentials > API Key
    5. Tạo file .env cùng thư mục với script này, thêm dòng:
       YOUTUBE_API_KEY=your_key_here
       (file .env đã được thêm vào .gitignore, không bị đẩy lên GitHub)

Cách chạy:
    python extract.py --keyword "data analyst" --max-results 100
"""

import os
import csv
import argparse
import datetime as dt
import time
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Đọc file .env (nếu có) và nạp vào biến môi trường của process hiện tại.
# File .env KHÔNG được commit lên Git (đã thêm vào .gitignore).
load_dotenv()

# ============================================================
# CONFIG - SỬA CÁC GIÁ TRỊ Ở ĐÂY, KHÔNG CẦN TRUYỀN THAM SỐ DÒNG LỆNH
# Nếu muốn chạy kiểu "python extract.py" trực tiếp trong VS Code / IDE
# thì chỉ cần sửa 4 dòng bên dưới rồi bấm Run.
# ============================================================
KEYWORD = "data analyst"          # từ khóa / chủ đề muốn theo dõi

_user_input = input("Hãy nhập số video bạn muốn lấy (Enter để dùng mặc định 100): ").strip()
MAX_RESULTS = int(_user_input) if _user_input else 100   # số lượng video muốn lấy

OUTPUT_DIR = "data/raw"           # thư mục lưu file CSV
# ============================================================

SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"
VIDEOS_URL = "https://www.googleapis.com/youtube/v3/videos"
CHANNELS_URL = "https://www.googleapis.com/youtube/v3/channels"

# cache subscriber_count theo channel_id, tránh gọi API trùng lặp cho cùng 1 kênh
CHANNEL_STATS_CACHE = {}


def get_channel_stats(channel_ids: list, api_key: str) -> dict:
    """Trả về dict {channel_id: subscriber_count}, chỉ gọi API cho channel_id chưa có trong cache."""
    to_fetch = [cid for cid in set(channel_ids) if cid not in CHANNEL_STATS_CACHE]

    for i in range(0, len(to_fetch), 50):  # API chỉ cho tối đa 50 id / request
        chunk = to_fetch[i : i + 50]
        params = {
            "part": "statistics",
            "id": ",".join(chunk),
            "key": api_key,
        }
        resp = requests.get(CHANNELS_URL, params=params)
        resp.raise_for_status()
        data = resp.json()

        for item in data.get("items", []):
            channel_id = item.get("id")
            stats = item.get("statistics", {})
            # hiddenSubscriberCount = True nghĩa là kênh đó ẩn số subscriber công khai
            if stats.get("hiddenSubscriberCount"):
                CHANNEL_STATS_CACHE[channel_id] = "hidden"
            else:
                CHANNEL_STATS_CACHE[channel_id] = stats.get("subscriberCount", 0)

        time.sleep(0.2)

    return CHANNEL_STATS_CACHE

# map category_id -> tên category (YouTube trả về id số, cần map sang tên)
CATEGORY_CACHE = {}


def get_api_key():
    api_key = os.getenv("YOUTUBE_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "Chưa có API key. Tạo file .env ở cùng thư mục với script, "
            "thêm dòng: YOUTUBE_API_KEY=your_key_here"
        )
    return api_key


def to_rfc3339(date_str: str, end_of_day: bool = False) -> str:
    """Chuyển 'YYYY-MM-DD' sang định dạng RFC3339 mà YouTube API yêu cầu."""
    time_part = "T23:59:59Z" if end_of_day else "T00:00:00Z"
    return f"{date_str}{time_part}"


def parse_duration(iso_duration: str) -> str:
    """Chuyển ISO 8601 duration (PT#H#M#S) sang dạng HH:MM:SS."""
    import re
    match = re.match(
        r"P(?:(?P<days>\d+)D)?T?(?:(?P<hours>\d+)H)?(?:(?P<minutes>\d+)M)?(?:(?P<seconds>\d+)S)?",
        iso_duration,
    )
    if not match:
        return "00:00:00"
    parts = match.groupdict()
    days = int(parts["days"] or 0)
    hours = int(parts["hours"] or 0) + days * 24
    minutes = int(parts["minutes"] or 0)
    seconds = int(parts["seconds"] or 0)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def get_category_map(api_key: str, region_code: str = "VN") -> dict:
    """Lấy map category_id -> category_name (cache lại để đỡ gọi API nhiều lần)."""
    if CATEGORY_CACHE:
        return CATEGORY_CACHE
    url = "https://www.googleapis.com/youtube/v3/videoCategories"
    params = {"part": "snippet", "regionCode": region_code, "key": api_key}
    resp = requests.get(url, params=params)
    resp.raise_for_status()
    for item in resp.json().get("items", []):
        CATEGORY_CACHE[item["id"]] = item["snippet"]["title"]
    return CATEGORY_CACHE


def search_video_ids(
    keyword: str,
    api_key: str,
    published_after: str = None,
    published_before: str = None,
    max_results: int = 50,
) -> list:
    """Search video theo keyword + khoảng thời gian, trả về list video_id."""
    video_ids = []
    page_token = None

    while len(video_ids) < max_results:
        params = {
            "part": "id",
            "q": keyword,
            "type": "video",
            "maxResults": min(50, max_results - len(video_ids)),
            "key": api_key,
            "order": "viewCount",  # ưu tiên video nhiều view lên trước (thay vì relevance)
        }
        if published_after:
            params["publishedAfter"] = to_rfc3339(published_after)
        if published_before:
            params["publishedBefore"] = to_rfc3339(published_before, end_of_day=True)
        if page_token:
            params["pageToken"] = page_token

        resp = requests.get(SEARCH_URL, params=params)
        resp.raise_for_status()
        data = resp.json()

        for item in data.get("items", []):
            video_ids.append(item["id"]["videoId"])

        page_token = data.get("nextPageToken")
        if not page_token:
            break
        time.sleep(0.2)  # tránh gọi API quá dồn dập

    return video_ids[:max_results]


def get_video_details(video_ids: list, api_key: str) -> list:
    """Gọi videos.list để lấy full metadata (snippet + statistics + contentDetails)."""
    records = []
    category_map = get_category_map(api_key)
    collection_date = dt.datetime.now().strftime("%Y-%m-%d")

    # API chỉ cho tối đa 50 id / request
    for i in range(0, len(video_ids), 50):
        chunk = video_ids[i : i + 50]
        params = {
            "part": "snippet,statistics,contentDetails",
            "id": ",".join(chunk),
            "key": api_key,
        }
        resp = requests.get(VIDEOS_URL, params=params)
        resp.raise_for_status()
        data = resp.json()

        for item in data.get("items", []):
            snippet = item.get("snippet", {})
            stats = item.get("statistics", {})
            content = item.get("contentDetails", {})
            video_id = item.get("id")

            record = {
                "video_id": video_id,
                "video_title": snippet.get("title", ""),
                "channel_id": snippet.get("channelId", ""),
                "channel_name": snippet.get("channelTitle", ""),
                "publish_date": snippet.get("publishedAt", ""),
                "video_url": f"https://www.youtube.com/watch?v={video_id}",
                "duration": parse_duration(content.get("duration", "PT0S")),
                "view_count": int(stats.get("viewCount", 0) or 0),
                "like_count": int(stats.get("likeCount", 0) or 0),
                "comment_count": int(stats.get("commentCount", 0) or 0),
                "tags": "|".join(snippet.get("tags", [])),
                "category": category_map.get(snippet.get("categoryId", ""), "Unknown"),
                "thumbnail_url": snippet.get("thumbnails", {})
                .get("high", {})
                .get("url", ""),
                "collection_date": collection_date,
            }
            records.append(record)

        time.sleep(0.2)

    # Lấy subscriber_count cho tất cả channel_id xuất hiện trong records (gọi 1 lần theo batch,
    # không gọi lặp lại cho cùng 1 kênh dù kênh đó có nhiều video trong kết quả).
    channel_ids = [r["channel_id"] for r in records if r["channel_id"]]
    channel_stats = get_channel_stats(channel_ids, api_key)
    for r in records:
        sub_count = channel_stats.get(r["channel_id"], "Unknown")
        # Giữ nguyên "hidden"/"Unknown" dạng text vì đây không phải số thật,
        # còn lại convert sang int để Excel/CSV nhận đúng kiểu số.
        r["subscriber_count"] = sub_count if sub_count in ("hidden", "Unknown") else int(sub_count)

    return records


def save_to_csv(records: list, output_dir: str = "data/raw") -> str:
    os.makedirs(output_dir, exist_ok=True)
    today_str = dt.datetime.now().strftime("%Y%m%d")
    filepath = os.path.join(output_dir, f"raw_youtube_{today_str}.csv")

    fieldnames = [
        "video_id",
        "video_title",
        "channel_id",
        "channel_name",
        "subscriber_count",
        "publish_date",
        "video_url",
        "duration",
        "view_count",
        "like_count",
        "comment_count",
        "tags",
        "category",
        "thumbnail_url",
        "collection_date",
    ]

    try:
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
    except PermissionError:
        raise PermissionError(
            f"Không ghi được file '{filepath}' vì đang bị khóa. "
            f"Khả năng cao file này đang MỞ trong Excel/chương trình khác — hãy đóng lại rồi chạy lại script."
        )

    return filepath


def save_to_excel(records: list, output_dir: str = "data/raw") -> str:
    """Xuất ra file .xlsx, đọc trực tiếp bằng Excel không bị lỗi font/encoding như CSV."""
    os.makedirs(output_dir, exist_ok=True)
    today_str = dt.datetime.now().strftime("%Y%m%d")
    filepath = os.path.join(output_dir, f"raw_youtube_{today_str}.xlsx")

    fieldnames = [
        "video_id",
        "video_title",
        "channel_id",
        "channel_name",
        "subscriber_count",
        "publish_date",
        "video_url",
        "duration",
        "view_count",
        "like_count",
        "comment_count",
        "tags",
        "category",
        "thumbnail_url",
        "collection_date",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "video_statistics"

    # header
    ws.append(fieldnames)

    # data rows
    for record in records:
        ws.append([record.get(field, "") for field in fieldnames])

    # auto-fit độ rộng cột (ước lượng theo độ dài text dài nhất trong cột)
    for col_idx, field in enumerate(fieldnames, start=1):
        max_len = len(field)
        for record in records:
            value = str(record.get(field, ""))
            max_len = max(max_len, len(value))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    try:
        wb.save(filepath)
    except PermissionError:
        raise PermissionError(
            f"Không ghi được file '{filepath}' vì đang bị khóa. "
            f"Khả năng cao file này đang MỞ trong Excel — hãy đóng file lại rồi chạy lại script."
        )
    return filepath


def main():
    # Dùng argparse chỉ để cho phép override CONFIG ở trên nếu muốn chạy qua dòng lệnh.
    # Nếu không truyền gì cả (vd bấm Run trong IDE), script sẽ tự dùng giá trị trong CONFIG.
    parser = argparse.ArgumentParser(description="Kéo dữ liệu YouTube theo từ khóa, lọc video trong tháng hiện tại")
    parser.add_argument("--keyword", default=KEYWORD, help="Từ khóa / chủ đề muốn theo dõi")
    parser.add_argument("--max-results", type=int, default=MAX_RESULTS, help="Số lượng video muốn lấy")
    parser.add_argument("--output-dir", default=OUTPUT_DIR, help="Thư mục lưu file CSV output")
    args = parser.parse_args()

    api_key = get_api_key()

    # Tính khung thời gian ĐỘNG: luôn là "từ ngày 1 của tháng hiện tại -> hôm nay".
    # Chạy ngày nào trong tháng cũng tự động lấy đúng phạm vi tháng đó, không cần sửa tay.
    today = dt.datetime.now()
    published_after = today.replace(day=1).strftime("%Y-%m-%d")
    published_before = today.strftime("%Y-%m-%d")

    print(f"[1/3] Đang search video với keyword='{args.keyword}', "
          f"trong tháng {today.month}/{today.year} ({published_after} -> {published_before}) ...")
    video_ids = search_video_ids(
        keyword=args.keyword,
        api_key=api_key,
        published_after=published_after,
        published_before=published_before,
        max_results=args.max_results,
    )
    print(f"   -> Tìm được {len(video_ids)} video")

    print("[2/3] Đang lấy chi tiết metadata cho từng video ...")
    records = get_video_details(video_ids, api_key)

    # Sort lại theo view_count thật (từ videos.list, chính xác hơn số view ước lượng
    # lúc search), để video "lọt top view" nằm lên đầu danh sách.
    records.sort(key=lambda r: int(r.get("view_count", 0) or 0), reverse=True)

    print("[3/3] Đang lưu ra CSV và Excel ...")
    csv_path = save_to_csv(records, args.output_dir)
    xlsx_path = save_to_excel(records, args.output_dir)
    print(f"   -> Đã lưu {len(records)} dòng vào:\n      CSV : {csv_path}\n      Excel: {xlsx_path}")


if __name__ == "__main__":
    main()